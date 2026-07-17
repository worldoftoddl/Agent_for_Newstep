"""표 질의 도구 테스트 — 심은 워크북(가상 데이터)으로 등록·질의·안전장치를 검증한다.

이식 원본: awesome-llm-apps/For_me/langgraph_data_analysis_agent (DataStore).
"""

import pytest
from openpyxl import Workbook

from agent.tools.table import (
    _normalize_column,
    _validate_sql,
    excel_load_table,
    excel_query_table,
)

FILE = "표본_매출.xlsx"
SHEET = "매출"
RANGE = "A3:D7"  # 헤더(3행) + 데이터 4행 — 위쪽 제목(A1)은 표 밖


@pytest.fixture()
def table_dir(tmp_path, monkeypatch):
    """제목 행 아래에 소형 매출 표를 심은 워크북 (전부 가상 데이터)."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws["A1"] = "매출 상세 (단위: 원)"
    rows = [
        ["거래처", "지역", "금액(원)", "확인"],
        ["A상사", "서울", 1000, "✓"],
        ["B물산", "부산", 2000, None],
        ["C상사", "서울", 3000, "✓"],
        ["D유통", "부산", 4000, "N/A"],
    ]
    for i, row in enumerate(rows, start=3):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    wb.save(tmp_path / FILE)
    monkeypatch.setenv("WORKPAPERS_DIR", str(tmp_path))
    return tmp_path


def test_normalize_column_preserves_korean():
    assert _normalize_column("금액(원)", 3) == "금액_원"
    assert _normalize_column("  Net Amount ", 1) == "net_amount"
    assert _normalize_column(None, 2) == "column_2"
    assert _normalize_column("2024잔액", 4) == "col_2024잔액"


def test_load_table_schema(table_dir):
    out = excel_load_table.invoke(
        {"path": FILE, "sheet": SHEET, "cell_range": RANGE}
    )
    assert "4행 × 4열" in out
    assert '"금액_원"' in out and '원본 헤더 "금액(원)"' in out
    assert "BIGINT" in out  # 숫자 열 타입 유지
    assert "A상사" in out  # 샘플 포함


def test_query_aggregation(table_dir):
    out = excel_query_table.invoke(
        {
            "path": FILE,
            "sheet": SHEET,
            "cell_range": RANGE,
            "sql": 'SELECT "지역", SUM("금액_원") AS 합계 FROM data GROUP BY 1 ORDER BY 2 DESC',
        }
    )
    assert "SQL 결과" in out and "2행" in out
    # 부산 6000 > 서울 4000
    assert out.index("부산") < out.index("서울")
    assert "6000" in out and "4000" in out


def test_query_without_header(table_dir):
    out = excel_query_table.invoke(
        {
            "path": FILE,
            "sheet": SHEET,
            "cell_range": "A4:D7",  # 헤더 없이 데이터만
            "sql": "SELECT SUM(c) AS s FROM data",
            "has_header": False,
        }
    )
    assert "10000" in out


def test_mixed_column_downgraded_to_text(table_dir):
    # 확인 열은 ✓/None/"N/A" 혼합 — 문자열로 강등되어 질의 가능해야 한다
    out = excel_query_table.invoke(
        {
            "path": FILE,
            "sheet": SHEET,
            "cell_range": RANGE,
            "sql": 'SELECT COUNT(*) AS n FROM data WHERE "확인" = \'✓\'',
        }
    )
    assert "| 2 |" in out


@pytest.mark.parametrize(
    "sql",
    [
        "INSERT INTO data VALUES (1)",
        "DROP TABLE data",
        "SELECT 1; SELECT 2",
        "SELECT * FROM other_table",
        "SELECT * FROM read_csv('/etc/passwd')",
        "ATTACH '/tmp/x.db' AS x",
    ],
)
def test_unsafe_sql_rejected(table_dir, sql):
    out = excel_query_table.invoke(
        {"path": FILE, "sheet": SHEET, "cell_range": RANGE, "sql": sql}
    )
    assert out.startswith("오류:")


def test_validate_sql_allows_cte():
    safe = _validate_sql(
        'WITH t AS (SELECT "지역", "금액_원" FROM data) SELECT * FROM t'
    )
    assert safe.upper().startswith("WITH")


def test_result_row_cap(tmp_path, monkeypatch):
    wb = Workbook()
    ws = wb.active
    ws.title = "긴표"
    ws.append(["n"])
    for i in range(150):
        ws.append([i])
    wb.save(tmp_path / "긴표.xlsx")
    monkeypatch.setenv("WORKPAPERS_DIR", str(tmp_path))
    out = excel_query_table.invoke(
        {
            "path": "긴표.xlsx",
            "sheet": "긴표",
            "cell_range": "A1:A151",
            "sql": "SELECT * FROM data",
        }
    )
    assert "100행" in out and "절단" in out


def test_missing_file_and_empty_range(table_dir):
    out = excel_load_table.invoke(
        {"path": "없는파일.xlsx", "sheet": SHEET, "cell_range": RANGE}
    )
    assert out.startswith("오류:")
    out = excel_load_table.invoke(
        {"path": FILE, "sheet": SHEET, "cell_range": "H20:J25"}
    )
    assert "데이터 행이 없습니다" in out
