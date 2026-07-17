"""표 질의 도구 — 시트 범위를 격리 DuckDB 테이블로 등록해 읽기 전용 SQL로 집계한다.

awesome-llm-apps/For_me/langgraph_data_analysis_agent의 DataStore(sqlglot AST
검증 + in-memory DuckDB, 외부 접근 차단)를 이식했다. 원본의 plan/revise 그래프
루프는 채택하지 않음 — create_agent의 ReAct 루프가 "오류: …" 텍스트를 받아
SQL을 자가 수정한다 (excel.py 도구들과 동일한 방침).

안전 장치 (2중):
  1) sqlglot AST 검증 — 단일 SELECT/WITH만 허용, DDL/DML·관리 구문 차단,
     data 테이블 외 접근 차단, read_*/scan/glob 등 외부 접근 함수 차단
  2) DuckDB enable_external_access=false — 엔진 수준에서 파일·URL 접근 차단
"""

import re
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path

import duckdb
import pandas as pd
from langchain_core.tools import tool
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from sqlglot import exp, parse

from agent.tools.excel import _load, _md, _resolve, _sheet

TABLE = "data"
MAX_TABLE_CELLS = 100_000
MAX_RESULT_ROWS = 100
SAMPLE_ROWS = 5

# UPDATE 등 뮤테이션·관리 구문 노드 — sqlglot 버전에 없는 이름은 건너뛴다
_FORBIDDEN_NODES = tuple(
    node
    for name in (
        "Insert", "Update", "Delete", "Create", "Drop", "Alter", "Copy",
        "Command", "Attach", "Detach", "Pragma", "Set", "Use", "Merge",
        "Transaction",
    )
    if (node := getattr(exp, name, None)) is not None
)


def _normalize_column(name, index: int) -> str:
    """헤더를 SQL 식별자로 정규화 — 한글은 보존, 공백·특수문자는 _ 로."""
    text = "" if name is None else str(name)
    value = re.sub(r"[^\w]+", "_", text.strip()).strip("_").lower()
    if not value:
        value = f"column_{index}"
    if value[0].isdigit():
        value = f"col_{value}"
    return value


def _coerce_columns(frame: pd.DataFrame) -> pd.DataFrame:
    """열 단위 타입 통일 — 전부 숫자면 숫자, 전부 날짜면 날짜, 혼합이면 문자열.

    조서 표는 숫자 열에 틱마크·"N/A" 같은 문자가 섞이곤 하는데, 혼합 object
    열은 DuckDB 등록이 깨지므로 문자열로 강등해 값을 보존한다.
    """
    for col in frame.columns:
        series = frame[col]
        non_null = series.dropna()
        if non_null.empty:
            continue
        if non_null.map(
            lambda v: isinstance(v, (int, float)) and not isinstance(v, bool)
        ).all():
            frame[col] = pd.to_numeric(series)
        elif non_null.map(lambda v: isinstance(v, (datetime, date))).all():
            frame[col] = pd.to_datetime(series)
        else:
            # pandas 3은 문자열 열의 None을 NaN으로 바꾸므로 pd.isna로 판정한다
            # — v is None 판정이면 str(nan) = "nan" 문자열이 생긴다 (실측)
            frame[col] = series.map(lambda v: None if pd.isna(v) else str(v))
    return frame


class _TableStore:
    def __init__(self, frame: pd.DataFrame, column_map: list[tuple[str, str]]):
        self.frame = frame
        self.column_map = column_map  # [(정규화 열명, 원본 헤더), …]
        self.connection = duckdb.connect(
            database=":memory:", config={"enable_external_access": "false"}
        )
        self.connection.register(TABLE, frame)


@lru_cache(maxsize=16)
def _store_cached(
    path_str: str, mtime: float, sheet: str, cell_range: str, has_header: bool
) -> _TableStore:
    target = Path(path_str)
    ws = _sheet(_load(target, data_only=True), sheet)
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    if has_header:
        originals = [
            ws.cell(row=min_row, column=c).value for c in range(min_col, max_col + 1)
        ]
        data_start = min_row + 1
    else:
        originals = [get_column_letter(c) for c in range(min_col, max_col + 1)]
        data_start = min_row

    columns, seen = [], set()
    for index, original in enumerate(originals, start=1):
        base = _normalize_column(original, index)
        candidate, suffix = base, 2
        while candidate in seen:
            candidate, suffix = f"{base}_{suffix}", suffix + 1
        seen.add(candidate)
        columns.append(candidate)

    rows = []
    for r in range(data_start, max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(min_col, max_col + 1)]
        if any(v is not None for v in row):
            rows.append(row)
    if not rows:
        raise ValueError(
            f"오류: 범위 {cell_range}에 데이터 행이 없습니다. "
            f"excel_workbook_overview의 블록 ref로 범위를 다시 확인하세요."
        )

    frame = _coerce_columns(pd.DataFrame(rows, columns=columns))
    column_map = [(norm, "" if orig is None else str(orig)) for norm, orig in zip(columns, originals)]
    return _TableStore(frame, column_map)


def _store(path: str, sheet: str, cell_range: str, has_header: bool) -> tuple[Path, _TableStore]:
    target = _resolve(path)
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    count = (max_row - min_row + 1) * (max_col - min_col + 1)
    if count > MAX_TABLE_CELLS:
        raise ValueError(
            f"오류: 범위가 {count:,}셀로 상한({MAX_TABLE_CELLS:,}셀)을 초과합니다. "
            f"표 본체 범위만 지정하세요."
        )
    return target, _store_cached(
        str(target), target.stat().st_mtime, sheet, cell_range, has_header
    )


def _validate_sql(sql: str) -> str:
    """단일 읽기 전용 SELECT/WITH만 통과시킨다. 위반 시 ValueError(사용자용 메시지)."""
    statements = [s for s in parse(sql.strip(), read="duckdb") if s is not None]
    if len(statements) != 1:
        raise ValueError("오류: SQL은 정확히 1개의 문장이어야 합니다.")
    statement = statements[0]
    if not isinstance(statement, (exp.Select, exp.Union)) and not statement.find(exp.Select):
        raise ValueError("오류: SELECT 또는 WITH 질의만 허용됩니다.")
    if any(statement.find(node) for node in _FORBIDDEN_NODES):
        raise ValueError("오류: 데이터 변경·관리 구문(DDL/DML 등)은 허용되지 않습니다.")
    tables = {table.name.lower() for table in statement.find_all(exp.Table)}
    ctes = {cte.alias.lower() for cte in statement.find_all(exp.CTE)}
    unknown = tables - ctes - {TABLE}
    if unknown:
        raise ValueError(
            f"오류: 알 수 없는 테이블 {', '.join(sorted(unknown))} — "
            f"질의 가능한 테이블은 {TABLE} 하나입니다."
        )
    for function in statement.find_all(exp.Func):
        name = function.sql_name().lower()
        if name.startswith(("read_", "scan")) or name in {"glob", "httpfs"}:
            raise ValueError(f"오류: 외부 접근 함수는 허용되지 않습니다: {name}")
    return statement.sql(dialect="duckdb")


def _markdown(columns: list[str], rows: list[list]) -> str:
    lines = [
        "| " + " | ".join(_md(str(c)) for c in columns) + " |",
        "|" + "---|" * len(columns),
    ]
    for row in rows:
        lines.append(
            "| " + " | ".join(_md("" if v is None else str(v)) for v in row) + " |"
        )
    return "\n".join(lines)


@tool
def excel_load_table(path: str, sheet: str, cell_range: str, has_header: bool = True) -> str:
    """시트의 표 범위를 SQL 질의용 테이블(이름: data)로 등록하고 스키마를 반환한다.

    집계·필터·정렬·검산 질문은 excel_read_range를 반복하지 말고, 이 도구로
    표를 등록한 뒤 excel_query_table로 SQL 한 번에 계산한다. 반환된 열명과
    타입을 확인하고 나서 질의를 작성할 것.

    Args:
        path: 조서 폴더 기준 파일명
        sheet: 시트 이름
        cell_range: 표 범위 — 헤더 행 포함 (excel_workbook_overview의 블록 ref 사용 가능)
        has_header: 범위 첫 행이 열 제목이면 True(기본). False면 열명이 Excel 열 문자(a, b, …)
    """
    try:
        target, store = _store(path, sheet, cell_range, has_header)
    except ValueError as e:
        return str(e)

    described = {
        name: kind
        for name, kind, *_ in store.connection.execute(f"DESCRIBE {TABLE}").fetchall()
    }
    nulls = store.frame.isna().sum()
    lines = [
        f"테이블 등록: {target.name} [{sheet}] {cell_range} → {TABLE} "
        f"({len(store.frame):,}행 × {len(store.frame.columns)}열)",
        '열 목록 — SQL에는 아래 열명을 사용 (한글·특수문자 열명은 큰따옴표로 감싸기):',
    ]
    for norm, orig in store.column_map:
        origin = f' ← 원본 헤더 "{orig}"' if orig and orig != norm else ""
        lines.append(f'- "{norm}" ({described.get(norm, "?")}, null {int(nulls[norm])}건){origin}')
    sample = store.frame.head(SAMPLE_ROWS).astype(object)
    lines.append(f"샘플 (처음 {len(sample)}행):")
    lines.append(_markdown(list(sample.columns), sample.where(pd.notna(sample), None).values.tolist()))
    lines.append(
        f"질의 방법: excel_query_table에 동일한 path/sheet/cell_range와 "
        f'SQL(예: SELECT "열명", SUM("숫자열") FROM {TABLE} GROUP BY 1)을 전달'
    )
    return "\n".join(lines)


@tool
def excel_query_table(
    path: str, sheet: str, cell_range: str, sql: str, has_header: bool = True
) -> str:
    """등록된 표(테이블명: data)에 읽기 전용 DuckDB SQL을 실행한다.

    합계 검증·조건 추출·그룹별 집계·상위 N 등 계산 질문에 사용한다.
    먼저 excel_load_table로 같은 범위를 등록해 열명·타입을 확인할 것.
    단일 SELECT 또는 WITH 질의만 허용되며(변경 구문·외부 파일 접근 차단),
    결과는 최대 100행 반환한다.

    Args:
        path: 조서 폴더 기준 파일명
        sheet: 시트 이름
        cell_range: excel_load_table에 넘긴 것과 동일한 표 범위
        sql: 실행할 SQL — FROM data, 한글 열명은 큰따옴표
             (예: SELECT "계정과목", SUM("금액") FROM data GROUP BY 1 ORDER BY 2 DESC)
        has_header: excel_load_table에 넘긴 것과 동일한 값
    """
    try:
        target, store = _store(path, sheet, cell_range, has_header)
        safe_sql = _validate_sql(sql)
    except ValueError as e:
        return str(e)

    limited = f"SELECT * FROM ({safe_sql}) AS q LIMIT {MAX_RESULT_ROWS + 1}"
    try:
        cursor = store.connection.execute(limited)
        columns = [item[0] for item in cursor.description]
        rows = [list(row) for row in cursor.fetchall()]
    except Exception as exc:  # DuckDB 오류는 에이전트가 읽고 SQL을 수정한다
        return f"오류: SQL 실행 실패 — {exc}"

    truncated = len(rows) > MAX_RESULT_ROWS
    rows = rows[:MAX_RESULT_ROWS]
    note = f" (상한 {MAX_RESULT_ROWS}행으로 절단 — 집계·필터로 좁히세요)" if truncated else ""
    return (
        f"SQL 결과: {target.name} [{sheet}] {cell_range} — {len(rows)}행{note}\n"
        f"실행 SQL: {safe_sql}\n" + _markdown(columns, rows)
    )


TABLE_TOOLS = [excel_load_table, excel_query_table]
