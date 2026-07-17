"""조서 기계 증거 수집 계층 — reviewer·explainer 공용 (비LLM).

워크북 구조·수식 지도·주석·서명란·본문 블록을 기존 Excel 도구 함수로
수집한다. 전부 결정적이라 같은 파일이면 같은 증거가 나온다.
"""

from openpyxl.utils import get_column_letter

from agent.tools.excel import (
    _detect_blocks,
    _load,
    _resolve,
    excel_formula_map,
    excel_get_annotations,
    excel_read_range,
    excel_workbook_overview,
)

MAX_SHEETS = 6
MAX_BLOCK_CELLS = 400  # 시트당 본문 정독 상한 (excel_read_range 상한 500 미만)
MAX_EVIDENCE_CHARS = 28_000

SIGNOFF_KEYWORDS = (
    "작성자", "검토자", "작성일", "검토일", "서명", "확인자",
    "Preparer", "Reviewer", "Prepared by", "Reviewed by",
)


def scan_signoffs(target) -> str:
    """작성/검토 표지 셀을 찾아 같은 셀·오른쪽·아래 셀의 채움 여부를 판정한다."""
    wb = _load(target, data_only=True)
    lines = []
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                text = value.strip()
                # 긴 문장 속 우연 일치(예: "…검토자는 다음을 확인한다")는 표지가 아님
                if len(text) > 30 or not any(k in text for k in SIGNOFF_KEYWORDS):
                    continue
                inline = text.split(":", 1)[1].strip() if ":" in text else ""
                right = ws.cell(row=cell.row, column=cell.column + 1).value
                below = ws.cell(row=cell.row + 1, column=cell.column).value
                filled = (
                    inline
                    or (str(right).strip() if right is not None else "")
                    or (str(below).strip() if below is not None else "")
                )
                status = f"채움({filled[:20]})" if filled else "공란"
                lines.append(f"- {ws.title}!{cell.coordinate} \"{text}\" → {status}")
    if not lines:
        return "[서명란 스캔] 작성·검토 표지를 찾지 못함"
    return "[서명란 스캔] (표지 셀 → 같은 셀·오른쪽·아래 값 유무)\n" + "\n".join(lines[:40])


def _clip_block_ref(block: dict) -> str:
    """블록을 정독 상한(MAX_BLOCK_CELLS) 이내로 행을 잘라 ref로 만든다."""
    rows = max(1, min(block["rows"], MAX_BLOCK_CELLS // max(1, block["cols"])))
    end_row = block["first_row"] + rows - 1
    return (
        f"{get_column_letter(block['c1'])}{block['first_row']}:"
        f"{get_column_letter(block['c2'])}{end_row}"
    )


def collect_workpaper_evidence(path_name: str) -> tuple[str, list[str], list[str]]:
    """조서 판단에 필요한 기계 증거를 수집한다 (비LLM).

    (증거 텍스트, 점검한 시트, 생략한 시트)를 돌려준다 — 보고서가 점검
    범위를 결정적으로 표기할 수 있게. 서명란 스캔은 핵심 증거라 증거
    선두(개요 직후)에 둔다 — 뒤에 두면 MAX_EVIDENCE_CHARS 절단 시 가장
    먼저 잘려나간다.
    """
    target = _resolve(path_name)
    wb = _load(target, data_only=True)
    visible = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
    examined = [ws.title for ws in visible[:MAX_SHEETS]]
    skipped = [ws.title for ws in visible[MAX_SHEETS:]]
    parts = [excel_workbook_overview.func(path_name), scan_signoffs(target)]
    for ws in visible[:MAX_SHEETS]:
        parts.append(excel_formula_map.func(path_name, ws.title))
        parts.append(excel_get_annotations.func(path_name, ws.title))
        blocks = _detect_blocks(ws)
        if blocks:
            largest = max(blocks, key=lambda b: b["rows"] * b["cols"])
            parts.append(
                excel_read_range.func(path_name, ws.title, _clip_block_ref(largest))
            )
    if skipped:
        parts.append(
            f"(주의: 시트 {len(skipped)}개는 증거 수집에서 생략됨 — "
            f"상한 {MAX_SHEETS}개: {', '.join(skipped)})"
        )
    evidence = "\n\n".join(parts)
    if len(evidence) > MAX_EVIDENCE_CHARS:
        evidence = evidence[:MAX_EVIDENCE_CHARS] + "\n… (증거 절단 — 상한 초과)"
    return evidence, examined, skipped
