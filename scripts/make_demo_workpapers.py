"""데모용 가상 샘플 조서 생성 — data/workpapers/에 xlsx 3건을 만든다.

시나리오 커버 (task_list Phase 6):
  1. 완성 조서   — 데모조서_5300 현금및현금성자산 (주)한빛전자.xlsx
  2. 미완성 조서 — 데모조서_5400 매출채권 (주)한빛전자 (작성중).xlsx
  3. 범용 Excel  — 데모_부서별 예산집행 현황 2026.xlsx

모든 수치·회사명은 가상이다 (실제 피감사회사 데이터 아님 — 공개 배포 제약).
에이전트 도구가 읽는 단서를 일부러 심는다: 틱마크+범례, 색상·볼드 마킹,
검토 서명란, 셀 메모, 수식 tie-out, 기준서 참조 문구.

openpyxl은 수식 셀의 캐시 값을 쓰지 못하므로(값 모드에서 공란으로 보임),
저장 후 sheet XML에 <v>를 직접 주입해 실제 Excel 저장본과 같게 만든다.

실행: .venv/bin/python scripts/make_demo_workpapers.py
"""

import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUT_DIR = Path(__file__).resolve().parent.parent / "data" / "workpapers"

# ── 공통 스타일 ──────────────────────────────────────────────────────────
TITLE = Font(bold=True, size=14)
BOLD = Font(bold=True)
RED_BOLD = Font(bold=True, color="CC0000")
HEADER_FILL = PatternFill("solid", start_color="D9E1F2")  # 연한 파랑
DONE_FILL = PatternFill("solid", start_color="C6EFCE")  # 연한 초록 — 대사 완료
PEND_FILL = PatternFill("solid", start_color="FFF2CC")  # 연한 노랑 — 미완료
OVER_FILL = PatternFill("solid", start_color="F8CBAD")  # 연한 주황 — 초과
THIN = Side(style="thin", color="999999")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WON = "#,##0"
PCT = "0.0%"

DISCLAIMER = "※ 본 파일은 ExcelBrief 데모용 가상 데이터이며 실제 회사·감사업무와 무관합니다."


def _header_row(ws, row: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=text)
        c.font = BOLD
        c.fill = HEADER_FILL
        c.border = BOX
        c.alignment = Alignment(horizontal="center")


def _money(ws, coord: str) -> None:
    ws[coord].number_format = WON
    ws[coord].border = BOX


class FormulaCache:
    """수식 셀의 계산 결과를 기록해 두었다가 저장 후 XML에 주입한다."""

    def __init__(self) -> None:
        self.by_sheet: dict[str, dict[str, int | float]] = {}

    def set(self, ws, coord: str, formula: str, value: int | float):
        ws[coord] = formula
        self.by_sheet.setdefault(ws.title, {})[coord] = value
        return ws[coord]


_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def patch_cached_values(path: Path, wb: Workbook, cache: FormulaCache) -> None:
    """sheet XML의 수식 셀(<c><f/></c>)에 캐시 값 <v>를 덧붙인다."""
    ET.register_namespace("", _NS)
    sheet_xml = {
        ws.title: f"xl/worksheets/sheet{i}.xml"
        for i, ws in enumerate(wb.worksheets, start=1)
    }
    with zipfile.ZipFile(path) as z:
        items = {n: z.read(n) for n in z.namelist()}
    for title, coords in cache.by_sheet.items():
        name = sheet_xml[title]
        root = ET.fromstring(items[name])
        patched = 0
        for c in root.iter(f"{{{_NS}}}c"):
            coord = c.get("r")
            if coord in coords and c.find(f"{{{_NS}}}f") is not None:
                # openpyxl이 수식 셀에 빈 <v/>를 이미 써 두므로 재사용한다
                v = c.find(f"{{{_NS}}}v")
                if v is None:
                    v = ET.SubElement(c, f"{{{_NS}}}v")
                val = coords[coord]
                v.text = repr(val) if isinstance(val, float) else str(val)
                patched += 1
        assert patched == len(coords), f"{title}: {patched}/{len(coords)}건만 주입됨"
        items[name] = ET.tostring(root, xml_declaration=True, encoding="UTF-8")
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        for n, data in items.items():
            z.writestr(n, data)


# ── 1. 완성 조서: 5300 현금및현금성자산 ─────────────────────────────────
def build_cash_workpaper(fc: FormulaCache) -> Workbook:
    wb = Workbook()

    # -- 5300 Lead ----------------------------------------------------------
    ws = wb.active
    ws.title = "5300 Lead"
    ws["A1"] = "(주)한빛전자 — 현금및현금성자산 리드시트"
    ws["A1"].font = TITLE
    ws["A2"] = "조서번호: 5300  |  감사기준일: 2025-12-31  |  단위: 원"
    ws["A3"] = "작성: 김선우 (2026-01-15)    검토: 박지훈 (2026-01-22)"
    ws["A3"].font = BOLD
    ws["A3"].comment = Comment(
        "검토 완료. 은행조회 대사 및 시산표 대사 이상 없음 확인. — 박지훈", "박지훈"
    )

    _header_row(ws, 5, ["계정과목", "당기말", "전기말", "증감액", "증감률", "Ref", "틱"])
    rows = [
        ("보통예금", 1_842_500_000, 1_517_300_000, "5310"),
        ("정기예금(만기 3개월 이내)", 1_000_000_000, 800_000_000, "5310"),
        ("현금성자산(MMF)", 350_000_000, 420_000_000, "5310"),
        ("소액현금", 3_200_000, 3_000_000, "실사"),
    ]
    for i, (name, cy, py, ref) in enumerate(rows, start=6):
        ws.cell(row=i, column=1, value=name).border = BOX
        ws.cell(row=i, column=2, value=cy)
        ws.cell(row=i, column=3, value=py)
        fc.set(ws, f"D{i}", f"=B{i}-C{i}", cy - py)
        fc.set(ws, f"E{i}", f'=IF(C{i}=0,"n/a",D{i}/C{i})', (cy - py) / py)
        ws.cell(row=i, column=6, value=ref)
        tick = ws.cell(row=i, column=7, value="✓")
        tick.fill = DONE_FILL
        tick.alignment = Alignment(horizontal="center")
        for col in "BCD":
            _money(ws, f"{col}{i}")
        ws[f"E{i}"].number_format = PCT
        ws[f"E{i}"].border = BOX
        ws[f"F{i}"].border = BOX
        ws[f"G{i}"].border = BOX

    cy_total = sum(r[1] for r in rows)
    py_total = sum(r[2] for r in rows)
    ws["A10"] = "합계"
    ws["A10"].font = BOLD
    fc.set(ws, "B10", "=SUM(B6:B9)", cy_total)
    fc.set(ws, "C10", "=SUM(C6:C9)", py_total)
    fc.set(ws, "D10", "=B10-C10", cy_total - py_total)
    for coord in ("B10", "C10", "D10"):
        _money(ws, coord)
        ws[coord].font = BOLD

    ws["A12"] = "T/B 대사: 시산표(현금및현금성자산) 금액"
    ws["B12"] = cy_total  # 시산표 금액 — 장부와 일치하는 시나리오
    ws["A13"] = "차이"
    fc.set(ws, "B13", "=B10-B12", 0)
    _money(ws, "B12")
    _money(ws, "B13")
    tick = ws["C12"]
    tick.value = "T/B"
    tick.fill = DONE_FILL
    tick.alignment = Alignment(horizontal="center")

    ws["A15"] = (
        "수행절차: ① 은행조회서 발송·회신 대사(5310) ② 소액현금 실사 입회 "
        "③ 시산표 대사 ④ 만기 3개월 초과 예금의 현금성자산 분류 검토"
    )
    ws["A16"] = (
        f"결론: 현금및현금성자산 {cy_total:,}원은 중요성 관점에서 "
        "적정하게 표시되어 있음."
    )
    ws["A16"].font = BOLD
    ws["A18"] = (
        "근거: 감사기준서 330에 따른 실증절차로서 감사기준서 505의 "
        "적극적 조회 방식으로 외부조회를 수행함. 현금성자산 분류 기준은 "
        "K-IFRS 제1007호 '현금흐름표' 문단 6~7의 정의를 적용함."
    )
    ws["A20"] = DISCLAIMER
    ws["A20"].font = Font(size=9, color="808080")
    ws.column_dimensions["A"].width = 34
    for col in "BCD":
        ws.column_dimensions[col].width = 16

    # -- 5310 은행조회 --------------------------------------------------------
    ws2 = wb.create_sheet("5310 은행조회")
    ws2["A1"] = "은행조회서 회신 대사표"
    ws2["A1"].font = TITLE
    ws2["A2"] = "조회기준일: 2025-12-31  |  발송일: 2026-01-02  |  단위: 원"
    _header_row(ws2, 4, ["은행", "계좌구분", "장부금액", "회신금액", "차이", "회신일", "틱"])
    banks = [
        ("한빛은행", "보통예금", 1_204_800_000, "2026-01-08"),
        ("가온은행", "보통예금", 637_700_000, "2026-01-09"),
        ("한빛은행", "정기예금", 1_000_000_000, "2026-01-08"),
        ("미르증권", "MMF", 350_000_000, "2026-01-18"),
    ]
    for i, (bank, kind, amt, replied) in enumerate(banks, start=5):
        ws2.cell(row=i, column=1, value=bank).border = BOX
        ws2.cell(row=i, column=2, value=kind).border = BOX
        ws2.cell(row=i, column=3, value=amt)
        ws2.cell(row=i, column=4, value=amt)
        fc.set(ws2, f"E{i}", f"=C{i}-D{i}", 0)
        ws2.cell(row=i, column=6, value=replied).border = BOX
        tick = ws2.cell(row=i, column=7, value="✓")
        tick.fill = DONE_FILL
        tick.alignment = Alignment(horizontal="center")
        for col in "CDE":
            _money(ws2, f"{col}{i}")
    ws2["F8"].comment = Comment(
        "미르증권 1차 조회서 미회신으로 1/12 2차 발송, 1/18 회신 수령. — 김선우",
        "김선우",
    )
    ws2["A10"] = "차이 합계"
    ws2["A10"].font = BOLD
    fc.set(ws2, "E10", "=SUM(E5:E8)", 0)
    _money(ws2, "E10")
    ws2["E10"].font = BOLD
    ws2["A12"] = "조회 대상 4개 계좌 전건 회신 수령, 차이 없음 확인."
    ws2.column_dimensions["A"].width = 14
    for col in "CDE":
        ws2.column_dimensions[col].width = 16
    ws2.column_dimensions["F"].width = 12

    # -- 범례 ----------------------------------------------------------------
    ws3 = wb.create_sheet("범례")
    ws3["A1"] = "틱마크 범례"
    ws3["A1"].font = TITLE
    _header_row(ws3, 3, ["틱마크", "의미"])
    legend = [
        ("✓", "은행조회서 회신금액과 대사 확인"),
        ("T/B", "시산표 금액과 대사 확인"),
        ("실사", "현금 실사 입회로 확인"),
    ]
    for i, (mark, meaning) in enumerate(legend, start=4):
        ws3.cell(row=i, column=1, value=mark).alignment = Alignment(horizontal="center")
        ws3.cell(row=i, column=1).border = BOX
        ws3.cell(row=i, column=2, value=meaning).border = BOX
    ws3["A7"] = "초록 배경 셀 = 절차 완료 항목"
    ws3.column_dimensions["B"].width = 40
    return wb


# ── 2. 미완성 조서: 5400 매출채권 (작성중) ──────────────────────────────
def build_receivables_workpaper(fc: FormulaCache) -> Workbook:
    wb = Workbook()

    # -- 5400 Lead ----------------------------------------------------------
    ws = wb.active
    ws.title = "5400 Lead"
    ws["A1"] = "(주)한빛전자 — 매출채권 리드시트 (작성중)"
    ws["A1"].font = TITLE
    ws["A2"] = "조서번호: 5400  |  감사기준일: 2025-12-31  |  단위: 원"
    ws["A3"] = "작성: 김선우 (2026-01-20)    검토: (미완료)"
    ws["A3"].font = BOLD
    _header_row(ws, 5, ["계정과목", "당기말", "전기말", "증감액", "Ref", "틱"])
    rows = [
        ("외상매출금", 4_218_600_000, 3_654_200_000, "5410"),
        ("받을어음", 512_000_000, 486_500_000, "5410"),
        ("대손충당금", -164_300_000, -128_900_000, "5420"),
    ]
    for i, (name, cy, py, ref) in enumerate(rows, start=6):
        ws.cell(row=i, column=1, value=name).border = BOX
        ws.cell(row=i, column=2, value=cy)
        ws.cell(row=i, column=3, value=py)
        fc.set(ws, f"D{i}", f"=B{i}-C{i}", cy - py)
        ws.cell(row=i, column=5, value=ref).border = BOX
        ws.cell(row=i, column=6).border = BOX  # 틱마크 공란 — 절차 미완
        for col in "BCD":
            _money(ws, f"{col}{i}")
    cell = ws["B8"]
    cell.fill = PEND_FILL
    cell.comment = Comment(
        "잠정치. 5420 기대신용손실 산정 완료 후 확정 예정. — 김선우", "김선우"
    )
    cy_total = sum(r[1] for r in rows)
    py_total = sum(r[2] for r in rows)
    ws["A9"] = "합계(순액)"
    ws["A9"].font = BOLD
    fc.set(ws, "B9", "=SUM(B6:B8)", cy_total)
    fc.set(ws, "C9", "=SUM(C6:C8)", py_total)
    fc.set(ws, "D9", "=B9-C9", cy_total - py_total)
    for coord in ("B9", "C9", "D9"):
        _money(ws, coord)
        ws[coord].font = BOLD

    ws["A11"] = (
        "계획된 절차: ① 거래처 채권조회(감사기준서 505) — 5410 진행중 "
        "② 대손충당금(기대신용손실, K-IFRS 제1109호) 재계산 — 5420 미착수 "
        "③ 기간귀속 테스트(매출 cut-off) — 미착수"
    )
    ws["A12"] = "결론: (작성 예정)"
    ws["A12"].font = RED_BOLD
    ws["A14"] = DISCLAIMER
    ws["A14"].font = Font(size=9, color="808080")
    ws.column_dimensions["A"].width = 30
    for col in "BCD":
        ws.column_dimensions[col].width = 16

    # -- 5410 조회서 관리 -----------------------------------------------------
    ws2 = wb.create_sheet("5410 조회서")
    ws2["A1"] = "매출채권 조회서 발송·회신 관리"
    ws2["A1"].font = TITLE
    ws2["A2"] = "표본: 잔액 상위 8개 거래처 (모집단의 72%)  |  발송일: 2026-01-05"
    _header_row(ws2, 4, ["거래처", "장부금액", "회신금액", "차이", "상태", "대체절차"])
    parties = [
        ("(주)세움디스플레이", 1_120_400_000, 1_120_400_000, "회신"),
        ("가람테크(주)", 764_100_000, 758_300_000, "회신"),
        ("(주)누리반도체", 583_800_000, 583_800_000, "회신"),
        ("한결전자부품(주)", 402_600_000, None, "회신대기"),
        ("(주)미래로", 297_500_000, None, "회신대기"),
        ("돌섬산업(주)", 244_900_000, None, "회신대기"),
        ("(주)바로물류", 189_300_000, None, "회신대기"),
        ("새빛정밀(주)", 151_200_000, None, "회신대기"),
    ]
    for i, (name, book, reply, status) in enumerate(parties, start=5):
        ws2.cell(row=i, column=1, value=name).border = BOX
        ws2.cell(row=i, column=2, value=book)
        _money(ws2, f"B{i}")
        if reply is not None:
            ws2.cell(row=i, column=3, value=reply)
            fc.set(ws2, f"D{i}", f"=B{i}-C{i}", book - reply)
            _money(ws2, f"C{i}")
            _money(ws2, f"D{i}")
        else:
            ws2.cell(row=i, column=3).border = BOX
            ws2.cell(row=i, column=4).border = BOX
        st = ws2.cell(row=i, column=5, value=status)
        st.border = BOX
        st.alignment = Alignment(horizontal="center")
        if status == "회신대기":
            st.fill = PEND_FILL
        ws2.cell(row=i, column=6).border = BOX  # 대체절차 전부 공란
    ws2["D6"].comment = Comment(
        "차이 5,800,000원 — 12/30 반품 미반영으로 추정. 원인조사 필요. — 김선우",
        "김선우",
    )
    ws2["A14"] = "미회신 5건: 2차 발송 예정. 최종 미회신 시 대체절차(후속 회수 확인 등) 수행 필요."
    ws2["A14"].font = RED_BOLD
    ws2.column_dimensions["A"].width = 22
    for col in "BCD":
        ws2.column_dimensions[col].width = 16
    ws2.column_dimensions["F"].width = 14

    # -- 5420 대손충당금 — 헤더만 있는 빈 시트 --------------------------------
    ws3 = wb.create_sheet("5420 대손충당금")
    ws3["A1"] = "대손충당금(기대신용손실) 산정"
    ws3["A1"].font = TITLE
    ws3["A3"] = "TODO: 채권 연령분석 데이터 입력 및 기대신용손실(K-IFRS 제1109호) 재계산 수행"
    ws3["A3"].font = RED_BOLD
    _header_row(ws3, 5, ["연령구간", "채권잔액", "손실율", "기대신용손실", "전기 손실율", "비고"])
    for i, bucket in enumerate(
        ["3개월 이내", "3~6개월", "6~12개월", "12개월 초과"], start=6
    ):
        ws3.cell(row=i, column=1, value=bucket).border = BOX
        for col in range(2, 7):
            ws3.cell(row=i, column=col).border = BOX  # 전부 공란
    ws3.column_dimensions["A"].width = 14
    for col in "BDE":
        ws3.column_dimensions[col].width = 16
    return wb


# ── 3. 범용 Excel: 부서별 예산집행 현황 ─────────────────────────────────
def build_generic_budget(fc: FormulaCache) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "예산총괄"
    ws["A1"] = "2026년 상반기 부서별 예산집행 현황"
    ws["A1"].font = TITLE
    ws["A2"] = "기준일: 2026-06-30  |  단위: 천원  |  작성: 경영지원팀"
    _header_row(ws, 4, ["부서", "연간예산", "상반기 집행액", "잔액", "집행률", "비고"])
    depts = [
        ("연구개발팀", 850_000, 512_300, ""),
        ("영업팀", 420_000, 391_800, "하반기 예산 증액 요청"),
        ("경영지원팀", 310_000, 148_200, ""),
        ("생산관리팀", 560_000, 233_400, ""),
        ("품질보증팀", 180_000, 92_700, ""),
        ("마케팅팀", 240_000, 226_100, "행사비 초과 집행 주의"),
    ]
    for i, (dept, budget, spent, note) in enumerate(depts, start=5):
        ws.cell(row=i, column=1, value=dept).border = BOX
        ws.cell(row=i, column=2, value=budget)
        ws.cell(row=i, column=3, value=spent)
        fc.set(ws, f"D{i}", f"=B{i}-C{i}", budget - spent)
        rate = fc.set(ws, f"E{i}", f"=C{i}/B{i}", spent / budget)
        rate.number_format = PCT
        rate.border = BOX
        if spent / budget > 0.85:
            rate.fill = OVER_FILL
        ws.cell(row=i, column=6, value=note).border = BOX
        for col in "BCD":
            _money(ws, f"{col}{i}")
    b_total = sum(d[1] for d in depts)
    s_total = sum(d[2] for d in depts)
    ws["A11"] = "합계"
    ws["A11"].font = BOLD
    fc.set(ws, "B11", "=SUM(B5:B10)", b_total)
    fc.set(ws, "C11", "=SUM(C5:C10)", s_total)
    fc.set(ws, "D11", "=B11-C11", b_total - s_total)
    fc.set(ws, "E11", "=C11/B11", s_total / b_total)
    ws["E11"].number_format = PCT
    for coord in ("B11", "C11", "D11"):
        _money(ws, coord)
        ws[coord].font = BOLD
    ws["A13"] = "주황 배경 = 집행률 85% 초과 부서 (하반기 예산 관리 필요)"
    ws["A15"] = DISCLAIMER
    ws["A15"].font = Font(size=9, color="808080")
    ws.column_dimensions["A"].width = 16
    for col in "BCD":
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["F"].width = 24

    ws2 = wb.create_sheet("월별집행")
    ws2["A1"] = "월별 집행 추이 (전사 합계, 천원)"
    ws2["A1"].font = TITLE
    _header_row(ws2, 3, ["월", "집행액", "누계"])
    monthly = [248_100, 262_400, 305_800, 271_900, 289_200, 227_100]
    cum = 0
    for i, amt in enumerate(monthly, start=4):
        cum += amt
        ws2.cell(row=i, column=1, value=f"2026-{i - 3:02d}").border = BOX
        ws2.cell(row=i, column=2, value=amt)
        fc.set(ws2, f"C{i}", f"=B{i}" if i == 4 else f"=C{i - 1}+B{i}", cum)
        _money(ws2, f"B{i}")
        _money(ws2, f"C{i}")
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14
    return wb


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    targets = {
        "데모조서_5300 현금및현금성자산 (주)한빛전자.xlsx": build_cash_workpaper,
        "데모조서_5400 매출채권 (주)한빛전자 (작성중).xlsx": build_receivables_workpaper,
        "데모_부서별 예산집행 현황 2026.xlsx": build_generic_budget,
    }
    for name, builder in targets.items():
        fc = FormulaCache()
        wb = builder(fc)
        path = OUT_DIR / name
        wb.save(path)
        patch_cached_values(path, wb, fc)
        n = sum(len(v) for v in fc.by_sheet.values())
        print(f"생성: {path.name} (수식 캐시 {n}건 주입)")


if __name__ == "__main__":
    main()
